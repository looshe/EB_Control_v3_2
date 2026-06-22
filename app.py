from __future__ import annotations

import csv
import io
import os
import sqlite3
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, Response, flash, g, jsonify, redirect, render_template, request, send_from_directory, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
UPLOAD_DIR = BASE_DIR / "uploads"
EXPORT_DIR = BASE_DIR / "exports"
DB_PATH = INSTANCE_DIR / "eb_control.db"

for folder in [INSTANCE_DIR, UPLOAD_DIR, EXPORT_DIR]:
    folder.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("EB_CONTROL_SECRET", "troque-essa-chave-depois")
app.config["MAX_CONTENT_LENGTH"] = 80 * 1024 * 1024

STATUS_LIST = ["Pendente", "Em andamento", "Resolvido"]
CRITICIDADE_LIST = ["Baixa", "Normal", "Alta", "Crítica"]
TIPOS_PADRAO = [
    "Sem identificação", "Praça não atendida", "Recusa do destinatário", "Coleta errada",
    "Erro de emissão", "Extravio", "Falta de emissão", "Sobra", "Avaria", "Erro", "Controle operacional"
]
TIPOS_ENVOLVIDO_PADRAO = ["Motorista", "Ajudante", "Conferente", "Filial", "Comercial", "Cliente", "Outro"]
FILIAIS_LIST = ["SJB", "NHO", "CRI", "OUT", "HBG"]
AGENDAMENTO_STATUS = ["Pendente", "Agendado", "Finalizado"]
REENTREGA_AUT_STATUS = ["Aguardando autorização", "Autorizada", "Negada"]
REENTREGA_AG_STATUS = ["Pendente com cliente", "Agendado", "Finalizado"]
DEFAULT_ADMIN_USER = "admin"
DEFAULT_ADMIN_PASSWORD = os.environ.get("EB_ADMIN_PASSWORD", "EB@Admin2026!")


def db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_: Exception | None = None) -> None:
    conn = g.pop("db", None)
    if conn is not None:
        conn.close()


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_str() -> str:
    return date.today().isoformat()


def proper_name(value: str | None) -> str:
    if not value:
        return ""
    value = " ".join(value.strip().split())
    keep_upper = {"NF", "CTE", "CT-E", "CTR", "LTDA", "ME", "S/A", "SA", "EPP", "SC", "PR", "RS"}
    words = []
    for word in value.split():
        raw = word.strip()
        up = raw.upper()
        if up in keep_upper or len(raw) <= 2 and up.isalpha():
            words.append(up)
        else:
            words.append(raw[:1].upper() + raw[1:].lower())
    return " ".join(words)


def norm(value: str | None) -> str:
    return proper_name(value).casefold()


def execute(sql: str, params: tuple = ()) -> sqlite3.Cursor:
    cur = db().execute(sql, params)
    db().commit()
    return cur


def init_db() -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            usuario TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL,
            ativo INTEGER NOT NULL DEFAULT 1,
            is_admin INTEGER NOT NULL DEFAULT 0,
            criado_em TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS pendencias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE,
            data_abertura TEXT NOT NULL,
            data_resolucao TEXT,
            cliente TEXT,
            cliente_norm TEXT,
            cidade TEXT,
            cidade_norm TEXT,
            filial_responsavel TEXT DEFAULT 'SJB',
            cpf_cnpj TEXT,
            endereco TEXT,
            chave_nfe TEXT,
            valor_nf TEXT,
            nf TEXT,
            cte_ctr TEXT,
            status TEXT NOT NULL DEFAULT 'Pendente',
            tipo TEXT,
            criticidade TEXT DEFAULT 'Normal',
            descricao TEXT,
            andamento TEXT,
            resolucao TEXT,
            criado_por INTEGER,
            atualizado_em TEXT,
            FOREIGN KEY(criado_por) REFERENCES usuarios(id)
        );
        CREATE TABLE IF NOT EXISTS envolvidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pendencia_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            nome TEXT NOT NULL,
            nome_norm TEXT NOT NULL,
            FOREIGN KEY(pendencia_id) REFERENCES pendencias(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS anexos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pendencia_id INTEGER NOT NULL,
            nome_original TEXT NOT NULL,
            nome_arquivo TEXT NOT NULL,
            enviado_por INTEGER,
            enviado_em TEXT NOT NULL,
            FOREIGN KEY(pendencia_id) REFERENCES pendencias(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS historico (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pendencia_id INTEGER,
            usuario_id INTEGER,
            acao TEXT NOT NULL,
            detalhes TEXT,
            criado_em TEXT NOT NULL,
            FOREIGN KEY(pendencia_id) REFERENCES pendencias(id) ON DELETE CASCADE,
            FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
        );
        CREATE TABLE IF NOT EXISTS opcoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categoria TEXT NOT NULL,
            valor TEXT NOT NULL,
            valor_norm TEXT NOT NULL,
            UNIQUE(categoria, valor_norm)
        );
        CREATE TABLE IF NOT EXISTS cadastros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categoria TEXT NOT NULL,
            valor TEXT NOT NULL,
            valor_norm TEXT NOT NULL,
            UNIQUE(categoria, valor_norm)
        );
        CREATE TABLE IF NOT EXISTS contatos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            funcao TEXT NOT NULL,
            whatsapp TEXT NOT NULL,
            criado_em TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS agendamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE,
            data_cadastro TEXT NOT NULL,
            cliente TEXT,
            cidade TEXT,
            cpf_cnpj TEXT,
            endereco TEXT,
            chave_nfe TEXT,
            valor_nf TEXT,
            nf TEXT,
            cte_ctr TEXT,
            data_solicitada TEXT,
            data_agendada TEXT,
            status TEXT NOT NULL DEFAULT 'Pendente',
            observacao TEXT,
            atualizado_em TEXT,
            criado_por INTEGER,
            FOREIGN KEY(criado_por) REFERENCES usuarios(id)
        );
        CREATE TABLE IF NOT EXISTS reentregas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE,
            data_importacao TEXT NOT NULL,
            data_ocorrencia TEXT,
            cliente TEXT,
            cliente_norm TEXT,
            cpf_cnpj TEXT,
            cidade TEXT,
            nf TEXT,
            cte_ctr TEXT,
            motivo TEXT,
            telefone TEXT,
            status_autorizacao TEXT NOT NULL DEFAULT 'Aguardando autorização',
            status_agendamento TEXT NOT NULL DEFAULT 'Pendente com cliente',
            data_reentrega TEXT,
            observacao TEXT,
            atualizado_em TEXT,
            criado_por INTEGER,
            FOREIGN KEY(criado_por) REFERENCES usuarios(id)
        );
        """
    )
    conn.commit()
    # Migrações simples para bancos antigos
    for table, columns in {
        'pendencias': [
            ('filial_responsavel', "TEXT DEFAULT 'SJB'"),
            ('cpf_cnpj', 'TEXT'),
            ('endereco', 'TEXT'),
            ('chave_nfe', 'TEXT'),
            ('valor_nf', 'TEXT'),
        ],
        'agendamentos': [
            ('cpf_cnpj', 'TEXT'),
            ('endereco', 'TEXT'),
            ('chave_nfe', 'TEXT'),
            ('valor_nf', 'TEXT'),
        ],
    }.items():
        for col, typ in columns:
            try:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {col} {typ}")
            except sqlite3.OperationalError:
                pass
    conn.commit()
    try:
        cur.execute("ALTER TABLE usuarios ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    conn.commit()
    # usuário administrador padrão
    admin_row = cur.execute("SELECT * FROM usuarios WHERE usuario=?", (DEFAULT_ADMIN_USER,)).fetchone()
    if admin_row is None:
        cur.execute(
            "INSERT INTO usuarios (nome, usuario, senha_hash, ativo, is_admin, criado_em) VALUES (?, ?, ?, 1, 1, ?)",
            ("Administrador", DEFAULT_ADMIN_USER, generate_password_hash(DEFAULT_ADMIN_PASSWORD), now_str()),
        )
    else:
        if check_password_hash(admin_row[3], "admin123"):
            cur.execute(
                "UPDATE usuarios SET senha_hash=?, is_admin=1, ativo=1 WHERE usuario=?",
                (generate_password_hash(DEFAULT_ADMIN_PASSWORD), DEFAULT_ADMIN_USER),
            )
        else:
            cur.execute("UPDATE usuarios SET is_admin=1, ativo=1 WHERE usuario=?", (DEFAULT_ADMIN_USER,))
    if cur.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0] == 0:
        cur.execute(
            "INSERT INTO usuarios (nome, usuario, senha_hash, ativo, is_admin, criado_em) VALUES (?, ?, ?, 1, 1, ?)",
            ("Administrador", DEFAULT_ADMIN_USER, generate_password_hash(DEFAULT_ADMIN_PASSWORD), now_str()),
        )
    for item in TIPOS_PADRAO:
        try:
            cur.execute("INSERT INTO opcoes (categoria, valor, valor_norm) VALUES (?, ?, ?)", ("tipo_pendencia", item, norm(item)))
        except sqlite3.IntegrityError:
            pass
    for item in TIPOS_ENVOLVIDO_PADRAO:
        try:
            cur.execute("INSERT INTO opcoes (categoria, valor, valor_norm) VALUES (?, ?, ?)", ("tipo_envolvido", item, norm(item)))
        except sqlite3.IntegrityError:
            pass
    conn.commit()
    # códigos faltantes
    rows = cur.execute("SELECT id FROM pendencias WHERE codigo IS NULL OR codigo='' ORDER BY id").fetchall()
    for row in rows:
        cur.execute("UPDATE pendencias SET codigo=? WHERE id=?", (f"OC-{row[0]:06d}", row[0]))
    conn.commit()
    conn.close()


init_db()


def current_user() -> sqlite3.Row | None:
    uid = session.get("user_id")
    if not uid:
        return None
    return db().execute("SELECT * FROM usuarios WHERE id=? AND ativo=1", (uid,)).fetchone()


@app.context_processor
def inject_globals() -> dict[str, Any]:
    return {
        "usuario_atual": current_user(),
        "is_admin_atual": bool(current_user() and (current_user()["is_admin"] or current_user()["usuario"] == DEFAULT_ADMIN_USER)),
        "status_list": STATUS_LIST,
        "criticidade_list": CRITICIDADE_LIST,
        "filiais_list": FILIAIS_LIST,
        "agendamento_status": AGENDAMENTO_STATUS,
        "reentrega_aut_status": REENTREGA_AUT_STATUS,
        "reentrega_ag_status": REENTREGA_AG_STATUS,
        "now_year": datetime.now().year,
    }


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = current_user()
        if not user or not (user["is_admin"] or user["usuario"] == DEFAULT_ADMIN_USER):
            flash("Acesso restrito ao administrador.", "erro")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def log_action(pendencia_id: int | None, acao: str, detalhes: str = "") -> None:
    execute(
        "INSERT INTO historico (pendencia_id, usuario_id, acao, detalhes, criado_em) VALUES (?, ?, ?, ?, ?)",
        (pendencia_id, session.get("user_id"), acao, detalhes, now_str()),
    )


def add_cadastro(categoria: str, valor: str | None) -> None:
    value = proper_name(valor)
    if not value:
        return
    try:
        execute("INSERT INTO cadastros (categoria, valor, valor_norm) VALUES (?, ?, ?)", (categoria, value, norm(value)))
    except sqlite3.IntegrityError:
        pass


def get_options(categoria: str) -> list[str]:
    return [r["valor"] for r in db().execute("SELECT valor FROM opcoes WHERE categoria=? ORDER BY valor", (categoria,)).fetchall()]


def get_cadastros(categoria: str) -> list[str]:
    return [r["valor"] for r in db().execute("SELECT valor FROM cadastros WHERE categoria=? ORDER BY valor", (categoria,)).fetchall()]


def fetch_pendencia(pid: int) -> sqlite3.Row:
    row = db().execute("SELECT * FROM pendencias WHERE id=?", (pid,)).fetchone()
    if not row:
        raise FileNotFoundError("Pendência não encontrada")
    return row


def save_envolvidos(pendencia_id: int, tipos: list[str], nomes: list[str]) -> None:
    execute("DELETE FROM envolvidos WHERE pendencia_id=?", (pendencia_id,))
    for tipo, nome in zip(tipos, nomes):
        tipo_clean = proper_name(tipo)
        nome_clean = proper_name(nome)
        if not tipo_clean or not nome_clean:
            continue
        execute(
            "INSERT INTO envolvidos (pendencia_id, tipo, nome, nome_norm) VALUES (?, ?, ?, ?)",
            (pendencia_id, tipo_clean, nome_clean, norm(nome_clean)),
        )
        add_cadastro("envolvido", nome_clean)
        add_cadastro(f"envolvido_{norm(tipo_clean)}", nome_clean)
        try:
            execute("INSERT INTO opcoes (categoria, valor, valor_norm) VALUES (?, ?, ?)", ("tipo_envolvido", tipo_clean, norm(tipo_clean)))
        except sqlite3.IntegrityError:
            pass


def handle_uploads(pendencia_id: int) -> None:
    files = request.files.getlist("anexos")
    target = UPLOAD_DIR / f"OC_{pendencia_id:06d}"
    target.mkdir(exist_ok=True)
    for file in files:
        if not file or not file.filename:
            continue
        original = file.filename
        safe = secure_filename(original)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        final_name = f"{stamp}_{safe}"
        file.save(target / final_name)
        execute(
            "INSERT INTO anexos (pendencia_id, nome_original, nome_arquivo, enviado_por, enviado_em) VALUES (?, ?, ?, ?, ?)",
            (pendencia_id, original, final_name, session.get("user_id"), now_str()),
        )
        log_action(pendencia_id, "Anexo adicionado", original)


def pendencia_payload() -> dict[str, Any]:
    status = request.form.get("status") or "Pendente"
    if status not in STATUS_LIST:
        status = "Pendente"
    criticidade = request.form.get("criticidade") or "Normal"
    if criticidade not in CRITICIDADE_LIST:
        criticidade = "Normal"
    return {
        "data_abertura": request.form.get("data_abertura") or today_str(),
        "cliente": proper_name(request.form.get("cliente")),
        "cidade": proper_name(request.form.get("cidade")),
        "filial_responsavel": (request.form.get("filial_responsavel") or "SJB").strip().upper(),
        "cpf_cnpj": (request.form.get("cpf_cnpj") or "").strip(),
        "endereco": proper_name(request.form.get("endereco")),
        "chave_nfe": (request.form.get("chave_nfe") or "").strip(),
        "valor_nf": (request.form.get("valor_nf") or "").strip(),
        "nf": (request.form.get("nf") or "").strip(),
        "cte_ctr": (request.form.get("cte_ctr") or "").strip(),
        "status": status,
        "tipo": proper_name(request.form.get("tipo")),
        "criticidade": criticidade,
        "descricao": request.form.get("descricao", "").strip(),
        "andamento": request.form.get("andamento", "").strip(),
        "resolucao": request.form.get("resolucao", "").strip(),
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = (request.form.get("usuario") or "").strip()
        senha = request.form.get("senha") or ""
        row = db().execute("SELECT * FROM usuarios WHERE usuario=? AND ativo=1", (usuario,)).fetchone()
        if row and check_password_hash(row["senha_hash"], senha):
            session.clear()
            session["user_id"] = row["id"]
            return redirect(url_for("dashboard"))
        flash("Usuário ou senha inválidos.", "erro")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/criar-usuario", methods=["GET", "POST"])
def criar_usuario_publico():
    flash("Criação pública de usuários está desativada. Solicite acesso ao administrador.", "erro")
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    conn = db()
    month = datetime.now().strftime("%Y-%m")
    counts = {
        "pendente": conn.execute("SELECT COUNT(*) c FROM pendencias WHERE status='Pendente'").fetchone()["c"],
        "andamento": conn.execute("SELECT COUNT(*) c FROM pendencias WHERE status='Em andamento'").fetchone()["c"],
        "resolvido": conn.execute("SELECT COUNT(*) c FROM pendencias WHERE status='Resolvido'").fetchone()["c"],
        "criticas": conn.execute("SELECT COUNT(*) c FROM pendencias WHERE criticidade='Crítica' AND status!='Resolvido'").fetchone()["c"],
        "mes": conn.execute("SELECT COUNT(*) c FROM pendencias WHERE substr(data_abertura,1,7)=?", (month,)).fetchone()["c"],
    }
    recentes = conn.execute("SELECT * FROM pendencias ORDER BY atualizado_em DESC, id DESC LIMIT 8").fetchall()
    criticas = conn.execute("SELECT * FROM pendencias WHERE criticidade='Crítica' AND status!='Resolvido' ORDER BY data_abertura ASC LIMIT 5").fetchall()
    por_tipo = conn.execute("SELECT COALESCE(tipo,'Sem tipo') label, COUNT(*) total FROM pendencias GROUP BY COALESCE(tipo,'Sem tipo') ORDER BY total DESC LIMIT 8").fetchall()
    por_status = conn.execute("SELECT status label, COUNT(*) total FROM pendencias GROUP BY status ORDER BY total DESC").fetchall()
    por_cliente = conn.execute("SELECT COALESCE(cliente,'Sem cliente') label, COUNT(*) total FROM pendencias GROUP BY COALESCE(cliente,'Sem cliente') ORDER BY total DESC LIMIT 6").fetchall()
    por_filial = conn.execute("SELECT COALESCE(filial_responsavel,'Sem filial') label, COUNT(*) total FROM pendencias GROUP BY COALESCE(filial_responsavel,'Sem filial') ORDER BY total DESC LIMIT 6").fetchall()
    por_mes = conn.execute("SELECT substr(data_abertura,1,7) label, COUNT(*) total FROM pendencias GROUP BY substr(data_abertura,1,7) ORDER BY label DESC LIMIT 6").fetchall()[::-1]
    ag_counts = {
        "ag_pendente": conn.execute("SELECT COUNT(*) c FROM agendamentos WHERE status='Pendente'").fetchone()["c"],
        "reen_aut": conn.execute("SELECT COUNT(*) c FROM reentregas WHERE status_autorizacao='Aguardando autorização'").fetchone()["c"],
        "sem_contato": conn.execute("SELECT COUNT(*) c FROM reentregas WHERE COALESCE(telefone,'')=''").fetchone()["c"],
    }
    return render_template("dashboard.html", counts=counts, recentes=recentes, criticas=criticas, por_tipo=por_tipo, por_status=por_status, por_cliente=por_cliente, por_filial=por_filial, por_mes=por_mes, ag_counts=ag_counts)


def list_query(status: str | None = None):
    q = (request.args.get("q") or "").strip()
    params = []
    where = []
    if status:
        where.append("p.status=?")
        params.append(status)
    if q:
        like = f"%{q}%"
        where.append("(p.codigo LIKE ? OR p.cliente LIKE ? OR p.cidade LIKE ? OR p.nf LIKE ? OR p.cte_ctr LIKE ? OR p.tipo LIKE ? OR p.filial_responsavel LIKE ? OR EXISTS (SELECT 1 FROM envolvidos e WHERE e.pendencia_id=p.id AND e.nome LIKE ?))")
        params.extend([like, like, like, like, like, like, like, like])
    sql = "SELECT p.* FROM pendencias p"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY p.id DESC"
    rows = db().execute(sql, params).fetchall()
    return rows, q


@app.route("/pendencias")
@login_required
def pendencias():
    status = request.args.get("status")
    if status not in STATUS_LIST:
        status = None
    rows, q = list_query(status)
    return render_template("pendencias.html", pendencias=rows, q=q, status_atual=status, titulo=status or "Todas as pendências")


@app.route("/pendencia/nova", methods=["GET", "POST"])
@login_required
def nova_pendencia():
    if request.method == "POST":
        data = pendencia_payload()
        xml_data = parse_nfe_xml_storage(request.files.get("xml_nfe"))
        for key in ["cliente", "cpf_cnpj", "cidade", "endereco", "nf", "chave_nfe", "valor_nf"]:
            if xml_data.get(key) and not data.get(key):
                data[key] = xml_data[key]
        if xml_data.get("data_emissao") and not data.get("data_abertura"):
            data["data_abertura"] = xml_data["data_emissao"]
        data_resolucao = now_str() if data["status"] == "Resolvido" else ""
        cur = execute(
            """
            INSERT INTO pendencias (data_abertura, data_resolucao, cliente, cliente_norm, cidade, cidade_norm, filial_responsavel, cpf_cnpj, endereco, chave_nfe, valor_nf, nf, cte_ctr, status, tipo, criticidade, descricao, andamento, resolucao, criado_por, atualizado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (data["data_abertura"], data_resolucao, data["cliente"], norm(data["cliente"]), data["cidade"], norm(data["cidade"]), data["filial_responsavel"], data["cpf_cnpj"], data["endereco"], data["chave_nfe"], data["valor_nf"], data["nf"], data["cte_ctr"], data["status"], data["tipo"], data["criticidade"], data["descricao"], data["andamento"], data["resolucao"], session.get("user_id"), now_str()),
        )
        pid = cur.lastrowid
        codigo = f"OC-{pid:06d}"
        execute("UPDATE pendencias SET codigo=? WHERE id=?", (codigo, pid))
        add_cadastro("cliente", data["cliente"])
        add_cadastro("cidade", data["cidade"])
        if data["tipo"]:
            try:
                execute("INSERT INTO opcoes (categoria, valor, valor_norm) VALUES (?, ?, ?)", ("tipo_pendencia", data["tipo"], norm(data["tipo"])))
            except sqlite3.IntegrityError:
                pass
        save_envolvidos(pid, request.form.getlist("envolvido_tipo[]"), request.form.getlist("envolvido_nome[]"))
        handle_uploads(pid)
        log_action(pid, "Ocorrência criada", codigo)
        flash("Pendência cadastrada com sucesso.", "ok")
        return redirect(url_for("detalhe_pendencia", pid=pid))
    return render_template("form.html", pendencia=None, envolvidos=[], anexos=[], historico=[], tipos=get_options("tipo_pendencia"), tipos_envolvido=get_options("tipo_envolvido"), clientes=get_cadastros("cliente"), cidades=get_cadastros("cidade"), nomes=get_cadastros("envolvido"))


@app.route("/pendencia/<int:pid>")
@login_required
def detalhe_pendencia(pid: int):
    p = fetch_pendencia(pid)
    envolvidos = db().execute("SELECT * FROM envolvidos WHERE pendencia_id=? ORDER BY id", (pid,)).fetchall()
    anexos = db().execute("SELECT * FROM anexos WHERE pendencia_id=? ORDER BY id DESC", (pid,)).fetchall()
    hist = db().execute("SELECT h.*, u.nome usuario_nome FROM historico h LEFT JOIN usuarios u ON u.id=h.usuario_id WHERE h.pendencia_id=? ORDER BY h.id DESC", (pid,)).fetchall()
    return render_template("detalhe.html", p=p, envolvidos=envolvidos, anexos=anexos, historico=hist)


@app.route("/pendencia/<int:pid>/editar", methods=["GET", "POST"])
@login_required
def editar_pendencia(pid: int):
    p = fetch_pendencia(pid)
    if request.method == "POST":
        data = pendencia_payload()
        xml_data = parse_nfe_xml_storage(request.files.get("xml_nfe"))
        for key in ["cliente", "cpf_cnpj", "cidade", "endereco", "nf", "chave_nfe", "valor_nf"]:
            if xml_data.get(key):
                data[key] = xml_data[key]
        old_status = p["status"]
        data_resolucao = p["data_resolucao"]
        if data["status"] == "Resolvido" and old_status != "Resolvido":
            data_resolucao = now_str()
        elif data["status"] != "Resolvido":
            data_resolucao = ""
        execute(
            """
            UPDATE pendencias SET data_abertura=?, data_resolucao=?, cliente=?, cliente_norm=?, cidade=?, cidade_norm=?, filial_responsavel=?, cpf_cnpj=?, endereco=?, chave_nfe=?, valor_nf=?, nf=?, cte_ctr=?, status=?, tipo=?, criticidade=?, descricao=?, andamento=?, resolucao=?, atualizado_em=? WHERE id=?
            """,
            (data["data_abertura"], data_resolucao, data["cliente"], norm(data["cliente"]), data["cidade"], norm(data["cidade"]), data["filial_responsavel"], data["cpf_cnpj"], data["endereco"], data["chave_nfe"], data["valor_nf"], data["nf"], data["cte_ctr"], data["status"], data["tipo"], data["criticidade"], data["descricao"], data["andamento"], data["resolucao"], now_str(), pid),
        )
        add_cadastro("cliente", data["cliente"])
        add_cadastro("cidade", data["cidade"])
        save_envolvidos(pid, request.form.getlist("envolvido_tipo[]"), request.form.getlist("envolvido_nome[]"))
        handle_uploads(pid)
        if data["status"] != old_status:
            log_action(pid, "Status alterado", f"{old_status} → {data['status']}")
        log_action(pid, "Ocorrência editada", data["codigo"] if "codigo" in data else p["codigo"])
        flash("Pendência atualizada.", "ok")
        return redirect(url_for("detalhe_pendencia", pid=pid))
    envolvidos = db().execute("SELECT * FROM envolvidos WHERE pendencia_id=? ORDER BY id", (pid,)).fetchall()
    anexos = db().execute("SELECT * FROM anexos WHERE pendencia_id=? ORDER BY id DESC", (pid,)).fetchall()
    hist = db().execute("SELECT h.*, u.nome usuario_nome FROM historico h LEFT JOIN usuarios u ON u.id=h.usuario_id WHERE h.pendencia_id=? ORDER BY h.id DESC", (pid,)).fetchall()
    return render_template("form.html", pendencia=p, envolvidos=envolvidos, anexos=anexos, historico=hist, tipos=get_options("tipo_pendencia"), tipos_envolvido=get_options("tipo_envolvido"), clientes=get_cadastros("cliente"), cidades=get_cadastros("cidade"), nomes=get_cadastros("envolvido"))


@app.route("/uploads/<int:pid>/<path:filename>")
@login_required
def download_anexo(pid: int, filename: str):
    return send_from_directory(UPLOAD_DIR / f"OC_{pid:06d}", filename, as_attachment=False)




def admin_return(default: str = "cadastros"):
    destino = request.args.get("next") or request.form.get("next") or default
    if destino == "usuarios":
        return redirect(url_for("usuarios_admin"))
    return redirect(url_for("cadastros"))


@app.route("/cadastros", methods=["GET", "POST"])
@login_required
@admin_required
def cadastros():
    if request.method == "POST":
        categoria = request.form.get("categoria") or "tipo_pendencia"
        valor = proper_name(request.form.get("valor"))
        if categoria not in ["tipo_pendencia", "tipo_envolvido"]:
            categoria = "tipo_pendencia"
        if valor:
            try:
                execute("INSERT INTO opcoes (categoria, valor, valor_norm) VALUES (?, ?, ?)", (categoria, valor, norm(valor)))
                flash("Opção adicionada.", "ok")
            except sqlite3.IntegrityError:
                flash("Essa opção já existe.", "erro")
        return redirect(url_for("cadastros"))
    tipos = db().execute("SELECT * FROM opcoes ORDER BY categoria, valor").fetchall()
    usuarios = db().execute("SELECT * FROM usuarios ORDER BY nome").fetchall()
    contatos = db().execute("SELECT * FROM contatos ORDER BY funcao, nome").fetchall()
    return render_template("cadastros.html", tipos=tipos, usuarios=usuarios, contatos=contatos)


@app.route("/usuarios")
@login_required
@admin_required
def usuarios_admin():
    usuarios = db().execute("SELECT * FROM usuarios ORDER BY nome").fetchall()
    return render_template("usuarios.html", usuarios=usuarios)


@app.route("/usuarios/novo", methods=["POST"])
@login_required
@admin_required
def novo_usuario():
    nome = proper_name(request.form.get("nome"))
    usuario = (request.form.get("usuario") or "").strip().lower()
    senha = request.form.get("senha") or ""
    confirmar = request.form.get("confirmar") or ""
    is_admin = 1 if request.form.get("is_admin") == "1" else 0

    if not nome or not usuario or not senha:
        flash("Preencha nome, usuário e senha.", "erro")
        return admin_return()
    if senha != confirmar:
        flash("As senhas não conferem.", "erro")
        return admin_return()
    if len(senha) < 6:
        flash("A senha precisa ter pelo menos 6 caracteres.", "erro")
        return admin_return()

    try:
        execute(
            "INSERT INTO usuarios (nome, usuario, senha_hash, ativo, is_admin, criado_em) VALUES (?, ?, ?, 1, ?, ?)",
            (nome, usuario, generate_password_hash(senha), is_admin, now_str()),
        )
        flash("Usuário criado com sucesso.", "ok")
    except sqlite3.IntegrityError:
        flash("Esse usuário já existe.", "erro")
    return admin_return()


def _admin_count() -> int:
    row = db().execute("SELECT COUNT(*) c FROM usuarios WHERE ativo=1 AND is_admin=1").fetchone()
    return int(row["c"] if row else 0)


@app.route("/usuarios/<int:uid>/senha", methods=["POST"])
@login_required
@admin_required
def alterar_senha_usuario(uid: int):
    senha = request.form.get("senha") or ""
    confirmar = request.form.get("confirmar") or ""
    if len(senha) < 6:
        flash("A nova senha precisa ter pelo menos 6 caracteres.", "erro")
        return admin_return()
    if senha != confirmar:
        flash("As senhas não conferem.", "erro")
        return admin_return()
    execute("UPDATE usuarios SET senha_hash=? WHERE id=?", (generate_password_hash(senha), uid))
    flash("Senha alterada com sucesso.", "ok")
    return admin_return()


@app.route("/usuarios/<int:uid>/status", methods=["POST"])
@login_required
@admin_required
def alterar_status_usuario(uid: int):
    user = db().execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not user:
        flash("Usuário não encontrado.", "erro")
        return admin_return()

    novo_status = 0 if user["ativo"] else 1
    if user["is_admin"] and user["ativo"] and novo_status == 0 and _admin_count() <= 1:
        flash("Não é permitido desativar o último administrador ativo.", "erro")
        return admin_return()

    execute("UPDATE usuarios SET ativo=? WHERE id=?", (novo_status, uid))
    flash("Usuário ativado." if novo_status else "Usuário desativado.", "ok")
    return admin_return()


@app.route("/usuarios/<int:uid>/admin", methods=["POST"])
@login_required
@admin_required
def alterar_admin_usuario(uid: int):
    user = db().execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not user:
        flash("Usuário não encontrado.", "erro")
        return admin_return()

    novo_admin = 0 if user["is_admin"] else 1
    if user["is_admin"] and novo_admin == 0 and user["ativo"] and _admin_count() <= 1:
        flash("Não é permitido remover o último administrador ativo.", "erro")
        return admin_return()

    execute("UPDATE usuarios SET is_admin=? WHERE id=?", (novo_admin, uid))
    flash("Permissão de administrador atualizada.", "ok")
    return admin_return()


@app.route("/usuarios/<int:uid>/excluir", methods=["POST"])
@login_required
@admin_required
def excluir_usuario(uid: int):
    user = db().execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not user:
        flash("Usuário não encontrado.", "erro")
        return admin_return()
    if session.get("user_id") == uid:
        flash("Você não pode excluir o usuário que está usando agora.", "erro")
        return admin_return()
    if user["is_admin"] and user["ativo"] and _admin_count() <= 1:
        flash("Não é permitido excluir o último administrador ativo.", "erro")
        return admin_return()

    execute("DELETE FROM usuarios WHERE id=?", (uid,))
    flash("Usuário excluído.", "ok")
    return admin_return()


@app.route("/historico")
@login_required
def historico_geral():
    rows = db().execute("SELECT h.*, u.nome usuario_nome, p.codigo FROM historico h LEFT JOIN usuarios u ON u.id=h.usuario_id LEFT JOIN pendencias p ON p.id=h.pendencia_id ORDER BY h.id DESC LIMIT 300").fetchall()
    return render_template("historico.html", historico=rows)


@app.route("/exportar")
@login_required
def exportar():
    tipo = request.args.get("tipo", "ocorrencias")
    if tipo == "envolvidos":
        output = []
        header = ["codigo_ocorrencia", "tipo_envolvido", "nome_envolvido", "tipo_pendencia", "status", "criticidade", "data_abertura", "data_resolucao", "cliente", "cidade", "filial_responsavel", "nf", "cte_ctr"]
        output.append(header)
        rows = db().execute(
            """
            SELECT p.*, e.tipo tipo_envolvido, e.nome nome_envolvido
            FROM envolvidos e JOIN pendencias p ON p.id=e.pendencia_id
            ORDER BY p.id, e.id
            """
        ).fetchall()
        for r in rows:
            output.append([r["codigo"], r["tipo_envolvido"], r["nome_envolvido"], r["tipo"], r["status"], r["criticidade"], r["data_abertura"], r["data_resolucao"], r["cliente"], r["cidade"], r["filial_responsavel"], r["nf"], r["cte_ctr"]])
        filename = "relatorio_envolvidos.csv"
    else:
        output = []
        header = ["codigo_ocorrencia", "data_abertura", "data_resolucao", "mes", "ano", "cliente", "cpf_cnpj", "cidade", "filial_responsavel", "nf", "cte_ctr", "chave_nfe", "valor_nf", "status", "tipo_pendencia", "criticidade", "envolvido_1_tipo", "envolvido_1_nome", "envolvido_2_tipo", "envolvido_2_nome", "envolvido_3_tipo", "envolvido_3_nome", "envolvidos_todos", "dias_em_aberto", "resolvido_sim_nao"]
        output.append(header)
        rows = db().execute("SELECT * FROM pendencias ORDER BY id").fetchall()
        for p in rows:
            ev = db().execute("SELECT * FROM envolvidos WHERE pendencia_id=? ORDER BY id", (p["id"],)).fetchall()
            todos = "; ".join([f"{e['tipo']}: {e['nome']}" for e in ev])
            slots = []
            for i in range(3):
                if i < len(ev):
                    slots.extend([ev[i]["tipo"], ev[i]["nome"]])
                else:
                    slots.extend(["", ""])
            try:
                d0 = datetime.fromisoformat(p["data_abertura"][:10])
                if p["data_resolucao"]:
                    d1 = datetime.fromisoformat(p["data_resolucao"][:10])
                else:
                    d1 = datetime.now()
                dias = max((d1 - d0).days, 0)
            except Exception:
                dias = ""
            mes = p["data_abertura"][:7] if p["data_abertura"] else ""
            ano = p["data_abertura"][:4] if p["data_abertura"] else ""
            output.append([p["codigo"], p["data_abertura"], p["data_resolucao"], mes, ano, p["cliente"], p["cpf_cnpj"], p["cidade"], p["filial_responsavel"], p["nf"], p["cte_ctr"], p["chave_nfe"], p["valor_nf"], p["status"], p["tipo"], p["criticidade"], *slots, todos, dias, "Sim" if p["status"] == "Resolvido" else "Não"])
        filename = "relatorio_ocorrencias.csv"
    def generate():
        import io
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";")
        for row in output:
            buffer.seek(0)
            buffer.truncate(0)
            writer.writerow(row)
            yield buffer.getvalue()
    return Response(generate(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename={filename}"})



def clean_phone(value: str | None) -> str:
    return "".join(ch for ch in (value or "") if ch.isdigit())


def read_uploaded_table(file) -> list[dict[str, str]]:
    """Lê CSV ou XLSX simples e devolve linhas por nome de coluna."""
    name = (file.filename or "").lower()
    rows: list[dict[str, str]] = []
    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            if not data:
                return []
            headers = [str(x or "").strip() for x in data[0]]
            for row in data[1:]:
                rows.append({headers[i]: str(row[i] or "").strip() for i in range(min(len(headers), len(row)))})
            return rows
        except Exception:
            return []
    raw = file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin1", errors="ignore")
    sample = text[:2000]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    import io
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [{str(k or "").strip(): str(v or "").strip() for k, v in row.items()} for row in reader]


def pick(row: dict[str, str], *names: str) -> str:
    def key_norm(x: str) -> str:
        return x.lower().replace("ç","c").replace("ã","a").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("-","").replace("_","").replace(" ","")
    mapped = {key_norm(k): v for k, v in row.items()}
    for name in names:
        kn = key_norm(name)
        if kn in mapped and mapped[kn]:
            return mapped[kn]
    for k, v in row.items():
        kk = key_norm(k)
        for name in names:
            if key_norm(name) in kk and v:
                return v
    return ""


def normalize_date_from_ssw(value: str | None) -> str:
    """Normaliza datas do SSW para YYYY-MM-DD quando possível."""
    raw = (value or "").strip()
    if not raw:
        return today_str()

    # Remove horário, se vier junto.
    raw_date = raw.split()[0].strip()

    formats = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d-%m-%y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(raw_date, fmt).date().isoformat()
        except ValueError:
            pass

    return today_str()


def only_digits(value: str | None) -> str:
    return "".join(ch for ch in (value or "") if ch.isdigit())


def parse_nfe_xml_storage(file) -> dict[str, str]:
    """Lê XML de NF-e ou ZIP contendo XML e retorna dados principais."""
    if not file or not getattr(file, "filename", ""):
        return {}

    filename = (file.filename or "").lower()
    if not (filename.endswith(".xml") or filename.endswith(".zip")):
        return {}

    try:
        raw = file.read()
        if not raw:
            return {}

        if filename.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                xml_files = [name for name in zf.namelist() if name.lower().endswith(".xml")]
                if not xml_files:
                    return {}
                raw = zf.read(xml_files[0])

        root = ET.fromstring(raw)
    except Exception:
        return {}

    def local(tag: str) -> str:
        return tag.split("}", 1)[-1] if "}" in tag else tag

    inf = None
    for el in root.iter():
        if local(el.tag) == "infNFe":
            inf = el
            break
    if inf is None:
        return {}

    def child(parent, name):
        if parent is None:
            return None
        for c in list(parent):
            if local(c.tag) == name:
                return c
        return None

    def text(parent, *names):
        if parent is None:
            return ""
        for c in list(parent):
            if local(c.tag) in names:
                return (c.text or "").strip()
        return ""

    ide = child(inf, "ide")
    dest = child(inf, "dest")
    ender = child(dest, "enderDest")
    total = child(inf, "total")
    icms = child(total, "ICMSTot")

    chave = (inf.attrib.get("Id") or "").replace("NFe", "").strip()
    cidade = text(ender, "xMun")
    uf = text(ender, "UF")
    cidade_final = f"{cidade} - {uf}" if cidade and uf else cidade
    endereco = ", ".join([x for x in [text(ender, "xLgr"), text(ender, "nro"), text(ender, "xBairro")] if x])
    emissao = text(ide, "dhEmi") or text(ide, "dEmi")
    if emissao:
        emissao = emissao[:10]

    return {
        "nf": text(ide, "nNF"),
        "chave_nfe": chave,
        "data_emissao": emissao,
        "cliente": proper_name(text(dest, "xNome")),
        "cpf_cnpj": text(dest, "CNPJ") or text(dest, "CPF"),
        "cidade": proper_name(cidade_final),
        "endereco": proper_name(endereco),
        "valor_nf": text(icms, "vNF"),
    }


@app.route("/api/ler-xml-nfe", methods=["POST"])
@login_required
def api_ler_xml_nfe():
    dados = parse_nfe_xml_storage(request.files.get("xml_nfe"))
    if not dados:
        return jsonify({"ok": False, "erro": "Não foi possível ler o XML da NF-e."}), 400
    return jsonify({"ok": True, "dados": dados})


@app.route("/api/opcoes/adicionar", methods=["POST"])
@login_required
def api_adicionar_opcao():
    categoria = (request.form.get("categoria") or "").strip()
    valor = proper_name(request.form.get("valor"))
    categorias_permitidas = {"tipo_pendencia", "tipo_envolvido"}

    if categoria not in categorias_permitidas:
        return jsonify({"ok": False, "erro": "Categoria inválida."}), 400
    if not valor:
        return jsonify({"ok": False, "erro": "Informe um valor."}), 400

    try:
        execute(
            "INSERT INTO opcoes (categoria, valor, valor_norm) VALUES (?, ?, ?)",
            (categoria, valor, norm(valor)),
        )
        criado = True
    except sqlite3.IntegrityError:
        criado = False

    return jsonify({"ok": True, "valor": valor, "criado": criado})


def agendamentos_query():
    q = (request.args.get("q") or "").strip()
    params = []
    where = []

    if q:
        like = f"%{q}%"
        where.append("(codigo LIKE ? OR cliente LIKE ? OR cidade LIKE ? OR nf LIKE ? OR cte_ctr LIKE ? OR status LIKE ?)")
        params = [like] * 6

    sql = "SELECT * FROM agendamentos"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"

    rows = db().execute(sql, params).fetchall()
    return rows, q

@app.route("/agendamentos", methods=["GET", "POST"])
@login_required
def agendamentos():
    if request.method == "POST":
        xml_data = parse_nfe_xml_storage(request.files.get("xml_nfe"))

        cliente = proper_name(request.form.get("cliente")) or xml_data.get("cliente", "")
        cidade = proper_name(request.form.get("cidade")) or xml_data.get("cidade", "")
        cpf_cnpj = (request.form.get("cpf_cnpj") or "").strip() or xml_data.get("cpf_cnpj", "")
        endereco = proper_name(request.form.get("endereco")) or xml_data.get("endereco", "")
        chave_nfe = (request.form.get("chave_nfe") or "").strip() or xml_data.get("chave_nfe", "")
        valor_nf = (request.form.get("valor_nf") or "").strip() or xml_data.get("valor_nf", "")
        nf = (request.form.get("nf") or "").strip() or xml_data.get("nf", "")
        cte = (request.form.get("cte_ctr") or "").strip()
        data_solicitada = request.form.get("data_solicitada") or today_str()
        data_agendada = request.form.get("data_agendada") or ""
        status = request.form.get("status") or "Pendente"

        if status not in AGENDAMENTO_STATUS:
            status = "Pendente"

        obs = request.form.get("observacao", "").strip()

        cur = execute("""
            INSERT INTO agendamentos
            (data_cadastro, cliente, cidade, cpf_cnpj, endereco, chave_nfe, valor_nf, nf, cte_ctr,
             data_solicitada, data_agendada, status, observacao, atualizado_em, criado_por)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            now_str(), cliente, cidade, cpf_cnpj, endereco, chave_nfe, valor_nf, nf, cte,
            data_solicitada, data_agendada, status, obs, now_str(), session.get("user_id")
        ))

        codigo = f"AG-{cur.lastrowid:06d}"
        execute("UPDATE agendamentos SET codigo=? WHERE id=?", (codigo, cur.lastrowid))

        add_cadastro("cliente", cliente)
        add_cadastro("cidade", cidade)

        flash("Agendamento salvo.", "ok")
        return redirect(url_for("agendamentos"))

    rows, q = agendamentos_query()
    hoje = today_str()

    counts = {
        "pendente": db().execute("SELECT COUNT(*) c FROM agendamentos WHERE status='Pendente'").fetchone()["c"],
        "agendado": db().execute("SELECT COUNT(*) c FROM agendamentos WHERE status='Agendado'").fetchone()["c"],
        "finalizado": db().execute("SELECT COUNT(*) c FROM agendamentos WHERE status='Finalizado'").fetchone()["c"],
        "atrasado": db().execute(
            "SELECT COUNT(*) c FROM agendamentos WHERE status!='Finalizado' AND data_agendada!='' AND data_agendada < ?",
            (hoje,)
        ).fetchone()["c"],
    }

    por_status = db().execute(
        "SELECT status label, COUNT(*) total FROM agendamentos GROUP BY status ORDER BY total DESC"
    ).fetchall()

    return render_template("agendamentos.html", rows=rows, q=q, counts=counts, por_status=por_status)


@app.route("/agendamentos/tabela")
@login_required
def agendamentos_tabela():
    rows, q = agendamentos_query()
    return render_template("agendamentos_tabela.html", rows=rows, q=q)


@app.route("/agendamentos/imprimir")
@login_required
def agendamentos_imprimir():
    rows, q = agendamentos_query()
    return render_template("agendamentos_print.html", rows=rows, q=q, now=now_str())


@app.route("/agendamentos/<int:aid>/excluir", methods=["POST"])
@login_required
def excluir_agendamento(aid: int):
    ag = db().execute("SELECT * FROM agendamentos WHERE id=?", (aid,)).fetchone()

    if not ag:
        flash("Agendamento não encontrado.", "erro")
        return redirect(url_for("agendamentos"))

    execute("DELETE FROM agendamentos WHERE id=?", (aid,))
    flash("Agendamento excluído.", "ok")
    return redirect(url_for("agendamentos"))
@app.route("/agendamentos/<int:aid>/status", methods=["POST"])
@login_required
def agendamento_status_update(aid:int):
    status=request.form.get("status") or "Pendente"
    data_agendada=request.form.get("data_agendada") or ""
    obs=request.form.get("observacao") or ""
    if status not in AGENDAMENTO_STATUS: status="Pendente"
    execute("UPDATE agendamentos SET status=?, data_agendada=?, observacao=?, atualizado_em=? WHERE id=?", (status, data_agendada, obs, now_str(), aid))
    flash("Agendamento atualizado.", "ok")
    return redirect(url_for("agendamentos"))


@app.route("/reentregas", methods=["GET", "POST"])
@login_required
def reentregas():
    if request.method == "POST":
        cliente=proper_name(request.form.get("cliente")); cidade=proper_name(request.form.get("cidade"))
        cpf=(request.form.get("cpf_cnpj") or "").strip(); nf=(request.form.get("nf") or "").strip(); cte=(request.form.get("cte_ctr") or "").strip()
        motivo=proper_name(request.form.get("motivo")) or "Cliente ausente"; telefone=(request.form.get("telefone") or "").strip()
        data_oc=request.form.get("data_ocorrencia") or today_str(); obs=request.form.get("observacao","").strip()
        cur=execute("""
            INSERT INTO reentregas (data_importacao, data_ocorrencia, cliente, cliente_norm, cpf_cnpj, cidade, nf, cte_ctr, motivo, telefone, observacao, atualizado_em, criado_por)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (now_str(), data_oc, cliente, norm(cliente), cpf, cidade, nf, cte, motivo, telefone, obs, now_str(), session.get("user_id")))
        codigo=f"RE-{cur.lastrowid:06d}"; execute("UPDATE reentregas SET codigo=? WHERE id=?", (codigo, cur.lastrowid))
        add_cadastro("cliente", cliente); add_cadastro("cidade", cidade)
        flash("Reentrega salva.", "ok")
        return redirect(url_for("reentregas"))
    q=(request.args.get("q") or "").strip(); filtro=request.args.get("filtro") or ""
    params=[]; where=[]
    if q:
        like=f"%{q}%"; where.append("(codigo LIKE ? OR cliente LIKE ? OR cidade LIKE ? OR nf LIKE ? OR cte_ctr LIKE ? OR cpf_cnpj LIKE ? OR motivo LIKE ?)"); params += [like]*7
    if filtro == "sem_contato": where.append("COALESCE(telefone,'')=''")
    if filtro == "aguardando_aut": where.append("status_autorizacao='Aguardando autorização'")
    if filtro == "aguardando_ag": where.append("status_autorizacao='Autorizada' AND status_agendamento='Pendente com cliente'")
    sql="SELECT * FROM reentregas" + (" WHERE "+" AND ".join(where) if where else "") + " ORDER BY id DESC"
    rows=db().execute(sql, params).fetchall()
    counts={
        "aut": db().execute("SELECT COUNT(*) c FROM reentregas WHERE status_autorizacao='Aguardando autorização'").fetchone()["c"],
        "ag": db().execute("SELECT COUNT(*) c FROM reentregas WHERE status_autorizacao='Autorizada' AND status_agendamento='Pendente com cliente'").fetchone()["c"],
        "sem": db().execute("SELECT COUNT(*) c FROM reentregas WHERE COALESCE(telefone,'')=''").fetchone()["c"],
        "final": db().execute("SELECT COUNT(*) c FROM reentregas WHERE status_agendamento='Finalizado'").fetchone()["c"],
    }
    return render_template("reentregas.html", rows=rows, q=q, filtro=filtro, counts=counts)


@app.route("/reentregas/importar", methods=["POST"])
@login_required
def importar_reentregas():
    file = request.files.get("arquivo")
    if not file or not file.filename:
        flash("Selecione um CSV ou XLSX do SSW.", "erro")
        return redirect(url_for("reentregas"))

    rows = read_uploaded_table(file)

    # Códigos do SSW que geram fluxo de reentrega:
    # 9  = destinatário ausente/fechado
    # 38 = destinatário impossibilitado de receber mercadoria
    codigos_reentrega = {"9", "38"}

    analisadas = 0
    importadas = 0
    ignoradas = 0
    duplicadas = 0
    erros = 0

    for row in rows:
        analisadas += 1

        cod_ocorrencia = only_digits(pick(row, "COD_OCORRENCIA", "COD OCORRENCIA", "CÓD OCORRÊNCIA", "CODIGO OCORRENCIA", "CÓDIGO OCORRÊNCIA"))
        if cod_ocorrencia not in codigos_reentrega:
            ignoradas += 1
            continue

        cliente = proper_name(pick(row, "DESTINATARIO", "DESTINATÁRIO", "CLIENTE", "NOME"))
        nf = pick(row, "NOTA_FISCAL", "NOTA FISCAL", "NF", "NUMERO NF", "NÚMERO NF")
        cte = pick(row, "CTRC", "CTE", "CT-E", "CTR", "CONHECIMENTO")
        descr_ocorrencia = proper_name(pick(row, "DESCR_OCORRENCIA", "DESCR OCORRENCIA", "DESCRIÇÃO OCORRÊNCIA", "DESCRICAO OCORRENCIA", "OCORRENCIA", "OCORRÊNCIA"))
        data_oc = normalize_date_from_ssw(pick(row, "DATA_OCORRENCIA", "DATA OCORRENCIA", "DATA OCORRÊNCIA", "DATA"))
        cidade = proper_name(pick(row, "CIDADE_ENTREGA", "CIDADE ENTREGA", "CIDADE", "MUNICIPIO", "MUNICÍPIO"))
        uf = (pick(row, "UF_ENTREGA", "UF ENTREGA", "UF") or "").strip().upper()
        cpf_cnpj = pick(row, "CNPJ_DESTINATARIO", "CNPJ DESTINATARIO", "CNPJ_DESTINATÁRIO", "CPF_CNPJ", "CPF/CNPJ", "DOCUMENTO")
        telefone = pick(row, "TELEFONE", "FONE", "WHATSAPP", "CELULAR", "CONTATO")

        if cidade and uf and uf not in cidade.upper():
            cidade = f"{cidade} - {uf}"

        if not cliente:
            erros += 1
            continue

        motivo = f"{cod_ocorrencia} - {descr_ocorrencia}" if descr_ocorrencia else f"Código {cod_ocorrencia}"

        # Evita importar a mesma reentrega várias vezes.
        # Prioriza NF + CT-e; se não houver CT-e, verifica pela NF.
        if nf:
            if cte:
                exists = db().execute(
                    """
                    SELECT id FROM reentregas
                    WHERE nf=? AND cte_ctr=? AND status_agendamento!='Finalizado'
                    LIMIT 1
                    """,
                    (nf, cte),
                ).fetchone()
            else:
                exists = db().execute(
                    """
                    SELECT id FROM reentregas
                    WHERE nf=? AND status_agendamento!='Finalizado'
                    LIMIT 1
                    """,
                    (nf,),
                ).fetchone()

            if exists:
                duplicadas += 1
                continue

        try:
            cur = execute(
                """
                INSERT INTO reentregas
                (data_importacao, data_ocorrencia, cliente, cliente_norm, cpf_cnpj, cidade, nf, cte_ctr, motivo, telefone,
                 status_autorizacao, status_agendamento, atualizado_em, criado_por)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    now_str(),
                    data_oc,
                    cliente,
                    norm(cliente),
                    cpf_cnpj,
                    cidade,
                    nf,
                    cte,
                    motivo,
                    telefone,
                    "Aguardando autorização",
                    "Pendente com cliente",
                    now_str(),
                    session.get("user_id"),
                ),
            )
            execute("UPDATE reentregas SET codigo=? WHERE id=?", (f"RE-{cur.lastrowid:06d}", cur.lastrowid))
            importadas += 1
            add_cadastro("cliente", cliente)
            add_cadastro("cidade", cidade)
        except Exception:
            erros += 1

    flash(
        f"Importação concluída: {analisadas} linhas analisadas, {importadas} reentregas importadas, "
        f"{ignoradas} ignoradas, {duplicadas} duplicadas e {erros} erro(s).",
        "ok" if importadas else "erro",
    )
    return redirect(url_for("reentregas"))
    rows=read_uploaded_table(file); total=0
    for row in rows:
        cliente=proper_name(pick(row,"cliente","destinatario","destinatário","nome"))
        if not cliente: continue
        cidade=proper_name(pick(row,"cidade","municipio","município"))
        cpf=pick(row,"cpf","cnpj","cpf/cnpj","documento")
        nf=pick(row,"nf","nota","nota fiscal","numero nf","número nf")
        cte=pick(row,"cte","ct-e","ctr","conhecimento")
        motivo=proper_name(pick(row,"motivo","ocorrencia","ocorrência","descricao","descrição")) or "Cliente ausente"
        telefone=pick(row,"telefone","fone","whatsapp","celular","contato")
        data_oc=pick(row,"data","data ocorrência","data ocorrencia") or today_str()
        cur=execute("""
            INSERT INTO reentregas (data_importacao, data_ocorrencia, cliente, cliente_norm, cpf_cnpj, cidade, nf, cte_ctr, motivo, telefone, atualizado_em, criado_por)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (now_str(), data_oc, cliente, norm(cliente), cpf, cidade, nf, cte, motivo, telefone, now_str(), session.get("user_id")))
        execute("UPDATE reentregas SET codigo=? WHERE id=?", (f"RE-{cur.lastrowid:06d}", cur.lastrowid)); total += 1
        add_cadastro("cliente", cliente); add_cadastro("cidade", cidade)
    flash(f"Importação concluída: {total} reentregas criadas.", "ok")
    return redirect(url_for("reentregas"))


@app.route("/reentregas/atualizar", methods=["POST"])
@login_required
def atualizar_reentrega():
    rid=int(request.form.get("id") or 0)
    aut=request.form.get("status_autorizacao") or "Aguardando autorização"
    ag=request.form.get("status_agendamento") or "Pendente com cliente"
    telefone=request.form.get("telefone") or ""
    data=request.form.get("data_reentrega") or ""
    obs=request.form.get("observacao") or ""
    if aut not in REENTREGA_AUT_STATUS: aut="Aguardando autorização"
    if ag not in REENTREGA_AG_STATUS: ag="Pendente com cliente"
    execute("UPDATE reentregas SET status_autorizacao=?, status_agendamento=?, telefone=?, data_reentrega=?, observacao=?, atualizado_em=? WHERE id=?", (aut, ag, telefone, data, obs, now_str(), rid))
    flash("Reentrega atualizada.", "ok")
    return redirect(url_for("reentregas"))


def contato_por_funcao(funcao: str):
    return db().execute("SELECT * FROM contatos WHERE funcao=? ORDER BY id DESC LIMIT 1", (funcao,)).fetchone()


@app.route("/whatsapp/reentregas", methods=["POST"])
@login_required
def whatsapp_reentregas():
    ids=[int(x) for x in request.form.getlist("ids") if x.isdigit()]
    if not ids:
        flash("Selecione pelo menos uma reentrega.", "erro"); return redirect(url_for("reentregas"))
    rows=db().execute(f"SELECT * FROM reentregas WHERE id IN ({','.join(['?']*len(ids))}) ORDER BY id", ids).fetchall()
    contato=contato_por_funcao("Autorização de Reentrega")
    if not contato:
        flash("Cadastre um contato com função 'Autorização de Reentrega' em Cadastros.", "erro"); return redirect(url_for("reentregas"))
    linhas=["Bom dia!", "", "Solicito autorização para reentrega dos clientes abaixo:", ""]
    for r in rows:
        linhas.append(f"• NF {r['nf'] or '-'} - {r['cliente'] or '-'} - {r['cidade'] or '-'}")
    linhas += ["", f"Total: {len(rows)} reentrega(s).", "Obrigado."]
    import urllib.parse
    phone=clean_phone(contato["whatsapp"]); text=urllib.parse.quote("\n".join(linhas))
    return redirect(f"https://wa.me/{phone}?text={text}")


@app.route("/whatsapp/sem-contato", methods=["POST"])
@login_required
def whatsapp_sem_contato():
    ids=[int(x) for x in request.form.getlist("ids") if x.isdigit()]
    if not ids:
        flash("Selecione pelo menos um cliente sem contato.", "erro"); return redirect(url_for("reentregas", filtro="sem_contato"))
    rows=db().execute(f"SELECT * FROM reentregas WHERE id IN ({','.join(['?']*len(ids))}) ORDER BY id", ids).fetchall()
    contato=contato_por_funcao("SAC Ortobom")
    if not contato:
        flash("Cadastre um contato com função 'SAC Ortobom' em Cadastros.", "erro"); return redirect(url_for("reentregas", filtro="sem_contato"))
    linhas=["Bom dia!", "", "Solicito contato atualizado dos clientes abaixo para agendamento de reentrega:", ""]
    for r in rows:
        linhas.append(f"• Cliente: {r['cliente'] or '-'}")
        linhas.append(f"  CPF/CNPJ: {r['cpf_cnpj'] or '-'}")
        linhas.append(f"  NF: {r['nf'] or '-'}")
        linhas.append("")
    linhas.append("Obrigado.")
    import urllib.parse
    phone=clean_phone(contato["whatsapp"]); text=urllib.parse.quote("\n".join(linhas))
    return redirect(f"https://wa.me/{phone}?text={text}")


@app.route("/contatos/novo", methods=["POST"])
@login_required
@admin_required
def novo_contato():
    nome=proper_name(request.form.get("nome")); funcao=(request.form.get("funcao") or "").strip(); whats=(request.form.get("whatsapp") or "").strip()
    if not nome or not funcao or not whats:
        flash("Preencha nome, função e WhatsApp.", "erro"); return redirect(url_for("cadastros"))
    execute("INSERT INTO contatos (nome, funcao, whatsapp, criado_em) VALUES (?, ?, ?, ?)", (nome, funcao, whats, now_str()))
    flash("Contato salvo.", "ok")
    return redirect(url_for("cadastros"))


@app.template_filter("envolvidos_preview")
def envolvidos_preview(pid: int) -> str:
    ev = db().execute("SELECT tipo, nome FROM envolvidos WHERE pendencia_id=? ORDER BY id", (pid,)).fetchall()
    if not ev:
        return "-"
    parts = [f"{e['tipo']}: {e['nome']}" for e in ev[:3]]
    if len(ev) > 3:
        parts.append(f"+{len(ev)-3}")
    return " | ".join(parts)


@app.template_filter("dias_aberto")
def dias_aberto(p: sqlite3.Row) -> int | str:
    try:
        d0 = datetime.fromisoformat(p["data_abertura"][:10])
        d1 = datetime.fromisoformat(p["data_resolucao"][:10]) if p["data_resolucao"] else datetime.now()
        return max((d1 - d0).days, 0)
    except Exception:
        return "-"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
